"""
Amazon ASIN 采集系统 v3 - Server
轻量级 FastAPI 服务器，适合 1C/2GB 低配部署
"""
import os
import io
import csv
import json
import re
import asyncio
import ipaddress
import logging
import time
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Any, Set
from urllib.parse import urlparse

import httpx
from fastapi import FastAPI, Request, UploadFile, File, Form, Query, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import openpyxl

from common import config
from common.database import Database

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ==================== 全局状态 ====================

db: Optional[Database] = None
_worker_registry: Dict[str, Dict] = {}
_WORKER_ID_RE = re.compile(r'^[\w\-]{1,64}$')

# 运行时设置
_runtime_settings: dict = {}
_settings_version: int = 0
_SETTINGS_FILE = os.path.join(config.PROJECT_DIR, "runtime_settings.json")

# 全局并发协调
_global_coordinator: Dict[str, Dict] = {}

# Worker 重启标记
_worker_restart_flags: Dict[str, bool] = {}


def _default_settings() -> dict:
    return {
        "zip_code": config.DEFAULT_ZIP_CODE,
        "max_retries": config.MAX_RETRIES,
        "request_timeout": config.REQUEST_TIMEOUT,
        "session_rotate_every": config.SESSION_ROTATE_EVERY,
        "proxy_url": config.PROXY_URL,
        "token_bucket_rate": config.TOKEN_BUCKET_RATE,
        "initial_concurrency": config.INITIAL_CONCURRENCY,
        "max_concurrency": config.MAX_CONCURRENCY,
        "min_concurrency": config.MIN_CONCURRENCY,
        "global_max_concurrency": config.GLOBAL_MAX_CONCURRENCY,
        "global_max_qps": config.GLOBAL_MAX_QPS,
        "adjust_interval": config.ADJUST_INTERVAL_S,
        "target_latency": config.TARGET_LATENCY_S,
        "max_latency": config.MAX_LATENCY_S,
        "target_success_rate": config.TARGET_SUCCESS_RATE,
        "min_success_rate": config.MIN_SUCCESS_RATE,
        "block_rate_threshold": config.BLOCK_RATE_THRESHOLD,
        "cooldown_after_block": 15,
        "proxy_bandwidth_mbps": config.PROXY_BANDWIDTH_MBPS,
        "screenshot_browsers": 1,
        "screenshot_pages_per_browser": 4,
        "auto_scrape_schedules": [],
        # 后台自动重试失败任务（AI 自动化场景：不需要人工点击重试）
        "auto_retry_failed_enabled": True,
        "auto_retry_cycles": 2,          # 最多重试多少轮（每轮会再走 MAX_RETRIES 次常规重试）
        "auto_retry_delay_minutes": 5,   # 任务失败后至少等这么久才会被自动重入队
    }


def _load_settings():
    global _runtime_settings, _settings_version
    _runtime_settings = _default_settings()
    if os.path.isfile(_SETTINGS_FILE):
        try:
            with open(_SETTINGS_FILE) as f:
                saved = json.load(f)
            _runtime_settings.update(saved)
        except Exception as e:
            logger.warning(f"加载设置失败: {e}")
    _settings_version = 0


# 敏感字段：不持久化到磁盘，避免凭据泄漏（从环境变量/.env 读取）
_SENSITIVE_SETTINGS_KEYS = {"proxy_url", "tunnel_proxy_url"}


def _save_settings():
    try:
        # 剥离敏感字段后再写文件；运行时内存里仍保留，供 worker 同步
        safe = {k: v for k, v in _runtime_settings.items() if k not in _SENSITIVE_SETTINGS_KEYS}
        with open(_SETTINGS_FILE, "w") as f:
            json.dump(safe, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"保存设置失败: {e}")


def _remove_screenshot_files(file_paths: list):
    """删除截图物理文件，忽略不存在的"""
    for fp in file_paths:
        if not fp:
            continue
        # file_path 可能是相对路径 (/static/screenshots/...) 或绝对路径
        if fp.startswith("/static/"):
            full = os.path.join(config.PROJECT_DIR, "server", fp.lstrip("/"))
        else:
            full = fp
        try:
            if os.path.isfile(full):
                os.remove(full)
        except OSError:
            pass


def _register_worker(worker_id: str, enable_screenshot: bool = None, ip: str = None):
    if not _WORKER_ID_RE.match(worker_id):
        return
    now = time.time()
    if worker_id not in _worker_registry:
        _worker_registry[worker_id] = {
            "worker_id": worker_id,
            "first_seen": now,
            "last_seen": now,
            "tasks_pulled": 0,
            "results_submitted": 0,
            "enable_screenshot": True,
            "ip": ip,
        }
    _worker_registry[worker_id]["last_seen"] = now
    if enable_screenshot is not None:
        _worker_registry[worker_id]["enable_screenshot"] = enable_screenshot
    if ip is not None:
        _worker_registry[worker_id]["ip"] = ip


# ==================== 生命周期 ====================

@asynccontextmanager
async def lifespan(app):
    global db, _callback_send_queue
    db = Database()
    await db.connect()
    _load_settings()
    os.makedirs(config.EXPORT_DIR, exist_ok=True)
    os.makedirs(config.SCREENSHOT_DIR, exist_ok=True)
    os.makedirs(_SCHEDULES_DIR, exist_ok=True)
    # 回调发送队列：maxsize 限制内存上限（高峰短期堆积也不会爆内存）
    _callback_send_queue = asyncio.Queue(maxsize=1000)
    logger.info("数据库初始化完成")
    asyncio.create_task(_timeout_task_loop())
    asyncio.create_task(_auto_scrape_scheduler())
    asyncio.create_task(_completion_watcher())
    asyncio.create_task(_callback_dispatcher())
    # 后台 WAL 维护（每 120s TRUNCATE checkpoint，防 WAL 顶 64MB checkpoint 饥饿）
    db.start_maintenance(checkpoint_interval=120)
    # 启动期 optimize 改为异步：服务先就绪、worker 先能拉任务，ANALYZE 后台慢慢做
    # （此前同步执行在 2.4GB 库上会阻塞启动 2~3 分钟）
    asyncio.create_task(db.run_startup_optimize())
    yield
    if db:
        await db.close()
    logger.info("服务器关闭")


app = FastAPI(title="Amazon Scraper v3", version="3.0.0", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=config.STATIC_DIR), name="static")
templates = Jinja2Templates(directory=config.TEMPLATE_DIR)


def _cst_filter(value):
    """Jinja 过滤器 `| cst`：把 UTC 时间字符串转中国时间（Asia/Shanghai, UTC+8）显示。
    服务器存 UTC（系统时区 Etc/UTC），前端统一显示中国时间。"""
    if not value:
        return ""
    try:
        dt = datetime.strptime(str(value)[:19], "%Y-%m-%d %H:%M:%S")
        return (dt + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return value


templates.env.filters["cst"] = _cst_filter

# 中国时间（UTC+8）：用于批次名/导出文件名/定时任务调度判定，与前端展示口径一致。
# 注意：数据库里存储的 created_at/updated_at 仍是 UTC（系统时区 Etc/UTC），由前端 +8 显示。
_CN_TZ_OFFSET = timedelta(hours=8)


def _cn_now():
    """当前中国时间（Asia/Shanghai, UTC+8）。仅用于命名与调度判定，不用于 DB 存储。"""
    return datetime.utcnow() + _CN_TZ_OFFSET


# ==================== 后台任务 ====================

async def _timeout_task_loop():
    """每 30s 心跳感知回收 + 硬超时兜底 + 清理离线 worker + WAL checkpoint"""
    _wal_checkpoint_counter = 0
    while True:
        await asyncio.sleep(30)
        try:
            # 识别死 Worker：2 分钟无心跳
            now = time.time()
            heartbeat_cutoff = now - 120
            dead_worker_ids = [
                wid for wid, w in _worker_registry.items()
                if w["last_seen"] < heartbeat_cutoff
            ]

            # 回收死 Worker 的 processing 任务 + 硬超时兜底（合成一条 SQL 防双重 epoch bump）
            await db.reclaim_dead_worker_tasks(dead_worker_ids)

            # 自动重试失败任务（AI 场景免人工点击）
            if _runtime_settings.get("auto_retry_failed_enabled", True):
                try:
                    await db.auto_retry_failed_tasks(
                        max_auto_cycles=int(_runtime_settings.get("auto_retry_cycles", 2)),
                        delay_minutes=int(_runtime_settings.get("auto_retry_delay_minutes", 5)),
                    )
                except Exception as e:
                    logger.warning(f"auto_retry_failed_tasks 异常: {e}")

            # 清理 5 分钟无心跳的 worker 注册信息
            offline_cutoff = now - 300
            truly_dead = [wid for wid, w in _worker_registry.items() if w["last_seen"] < offline_cutoff]
            for wid in truly_dead:
                del _worker_registry[wid]
                _global_coordinator.pop(wid, None)

            # 清理导出临时文件（超过 1 小时的 export_*.xlsx）
            try:
                export_dir = config.EXPORT_DIR
                if os.path.isdir(export_dir):
                    for fname in os.listdir(export_dir):
                        if fname.startswith("export_") and fname.endswith(".xlsx"):
                            fpath = os.path.join(export_dir, fname)
                            if now - os.path.getmtime(fpath) > 3600:
                                os.remove(fpath)
            except Exception:
                pass

            # 兜底清理 openpyxl 泄漏的 /tmp 临时目录（超过 30 分钟未动）
            # 正常 wb.close() 会清掉；异常路径可能漏清，累积会吃满磁盘
            try:
                tmp_root = "/tmp"
                if os.path.isdir(tmp_root):
                    for fname in os.listdir(tmp_root):
                        if not fname.startswith("openpyxl."):
                            continue
                        fpath = os.path.join(tmp_root, fname)
                        try:
                            if now - os.path.getmtime(fpath) > 1800:
                                if os.path.isdir(fpath):
                                    import shutil as _sh
                                    _sh.rmtree(fpath, ignore_errors=True)
                                else:
                                    os.remove(fpath)
                                logger.info(f"清理 openpyxl 泄漏目录: {fname}")
                        except Exception:
                            pass
            except Exception:
                pass

            # 每 10 次循环（约 5 分钟）做一次 WAL checkpoint —— 按需 TRUNCATE 策略
            # 经 30k recon 实测：固定 TRUNCATE 会触发 ~200-400ms 的 commit 阻塞抖动，
            # 影响 worker pull_tasks 和 accept_results_batch（max hold 407ms）。
            # 新策略：
            #   默认走 PASSIVE（非阻塞，能 checkpoint 多少算多少）
            #   仅当 WAL 文件超过阈值才主动 TRUNCATE（兜底，几小时一次）
            # WAL 自身有 PRAGMA wal_autocheckpoint=1000 + journal_size_limit=64MB 兜底
            _wal_checkpoint_counter += 1
            if _wal_checkpoint_counter >= 10:
                _wal_checkpoint_counter = 0
                wal_size = 0
                try:
                    wal_path = config.DB_PATH + "-wal"
                    if os.path.exists(wal_path):
                        wal_size = os.path.getsize(wal_path)
                except Exception:
                    pass
                wal_mb = wal_size / 1024 / 1024
                # 阈值：WAL > 128MB 才主动 TRUNCATE；否则 PASSIVE 不阻塞 writer
                mode = "TRUNCATE" if wal_size > 128 * 1024 * 1024 else "PASSIVE"
                try:
                    res = await db.wal_checkpoint(mode)
                    if res:
                        logger.debug(
                            f"WAL checkpoint({mode}) wal={wal_mb:.1f}MB "
                            f"busy={res[0]} log={res[1]} checkpointed={res[2]}"
                        )
                except Exception as e:
                    logger.warning(f"WAL checkpoint 失败: {e}")
                # 分档告警
                if wal_size > 500 * 1024 * 1024:
                    logger.error(f"WAL 文件过大: {wal_mb:.0f}MB（>500MB）")
                elif wal_size > 200 * 1024 * 1024:
                    logger.warning(f"WAL 文件偏大: {wal_mb:.0f}MB（>200MB）")

            # 兜底扫描：把 DB 中到期可重试的 callback 重新入队
            # （服务重启、内存队列丢失、_completion_watcher 漏检测 等场景的安全网）
            try:
                if _callback_send_queue is not None:
                    now_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                    due = await db.list_callback_due(now_str, limit=50)
                    for row in due:
                        try:
                            _callback_send_queue.put_nowait(row["id"])
                        except asyncio.QueueFull:
                            break  # 队列满了下次再说
            except Exception as e:
                logger.warning(f"callback 兜底扫描异常: {e}")

            # 兜底完成检测：扫描长期 running 的批次（防止入队事件丢失）
            # 只扫最近活动的 batches，避免全表查询
            try:
                async with db._db.execute(
                    """SELECT id FROM batches
                       WHERE status='running'
                       ORDER BY updated_at DESC
                       LIMIT 30"""
                ) as c:
                    rows = await c.fetchall()
                for r in rows:
                    _completion_check_set.add(r[0])
            except Exception as e:
                logger.warning(f"completion 兜底扫描异常: {e}")
        except Exception as e:
            logger.error(f"超时任务循环异常: {e}")


async def _completion_watcher():
    """完成检测协程：消费 _completion_check_set，找出已完成的批次。

    特点：
    - 完全脱离 worker 写入热路径
    - 每 _COMPLETION_WATCHER_INTERVAL 秒醒来一次，批量处理
    - 单次最多处理 100 个 batch_id，避免长时间占用事件循环
    - 任何异常只 log，不中断协程
    """
    logger.info("✅ 完成检测协程启动")
    while True:
        await asyncio.sleep(_COMPLETION_WATCHER_INTERVAL)
        try:
            # 一次性取走集合（用 set 自身的"原子读取"——单线程 asyncio 下安全）
            if not _completion_check_set:
                continue
            batch_ids = list(_completion_check_set)[:100]
            for bid in batch_ids:
                _completion_check_set.discard(bid)

            for bid in batch_ids:
                try:
                    # 快速 SQL：判断 task + screenshot 是否全部终态
                    snap = await db.get_batch_completion_status(bid)
                    if not snap["all_terminal"]:
                        continue
                    # 变体自动展开（仅 expand_variants 批次）：本轮全部终态后，把已采 ASIN
                    # 的同族变体入队【同一批次】继续采。新增 >0 则本批未真正完成，下轮再查。
                    added = await db.expand_batch_variants(bid)
                    if added > 0:
                        rnd = _expand_rounds.get(bid, 0) + 1
                        _expand_rounds[bid] = rnd
                        if rnd >= 2:
                            logger.warning(
                                f"🧬 批次 {bid} 变体展开第 {rnd} 轮仍新增 {added} 个任务："
                                f"家族列表可能不一致，请留意（去重保证仍会收敛）")
                        else:
                            logger.info(f"🧬 批次 {bid} 变体展开：新增 {added} 个同族任务，继续采集")
                        _completion_check_set.add(bid)  # 等这轮采完再复查
                        continue
                    _expand_rounds.pop(bid, None)
                    # 全部终态且无新变体 → 状态机转移（幂等）
                    changed = await db.mark_batch_completed(bid)
                    if changed:
                        logger.info(f"✅ 批次 {bid} 已完成，入队回调")
                        # 入队 callback dispatcher（仅当批次配置了回调）
                        if _callback_send_queue is not None:
                            try:
                                _callback_send_queue.put_nowait(bid)
                            except asyncio.QueueFull:
                                logger.warning("⚠️ callback 发送队列已满，依赖兜底扫描重发")
                except Exception as e:
                    logger.warning(f"完成检测单批次异常 batch_id={bid}: {e}")
        except Exception as e:
            logger.error(f"完成检测协程异常: {e}")


async def _callback_dispatcher():
    """回调发送协程：串行消费 _callback_send_queue + DB 待发条目，POST 到调用方。

    串行（而非并发）的原因：
    - 调用方接收端通常单实例，并发推送反而堆积
    - 串行简单可控，单次 10s 超时不会拖累整体
    - 失败按退避表延迟，写回 DB 等待下次扫描

    防卡死：
    - asyncio.wait_for(post, timeout=10) 强制中断慢响应
    - 重试上限 5 次后 callback_status='failed'，永不再处理
    """
    logger.info("✅ 回调发送协程启动")
    # 独立 httpx client（不与采集相关 HTTP 复用）
    async with httpx.AsyncClient(timeout=_CALLBACK_HTTP_TIMEOUT,
                                  follow_redirects=False) as client:
        while True:
            try:
                # 阻塞等下一个 batch_id（队列空时不空转）
                batch_id = await _callback_send_queue.get()
            except Exception as e:
                logger.error(f"回调队列读取异常: {e}")
                await asyncio.sleep(1)
                continue
            try:
                await _send_one_callback(client, batch_id)
            except Exception as e:
                logger.error(f"回调发送异常 batch_id={batch_id}: {e}")


async def _send_one_callback(client: httpx.AsyncClient, batch_id: int):
    """发送单个 batch 的回调。失败按退避表写回 DB。"""
    # 读取 batch 当前状态（要求 callback_status='pending' 才发；其他状态可能已 sent/failed/取消）
    async with db._db.execute(
        """SELECT id, name, external_id, callback_url, callback_attempts,
                  callback_status, completed_at, status
           FROM batches WHERE id=?""",
        (batch_id,)
    ) as c:
        row = await c.fetchone()
    if not row:
        return
    batch = dict(row)
    if batch.get("callback_status") != "pending":
        return  # 已 sent / failed / disabled / 无回调
    callback_url = batch.get("callback_url")
    if not callback_url:
        return
    # 防御：完成检测协程可能比兜底扫描慢一拍，导致 callback_status='pending'
    # 但 status 还是 'running'。此时跳过，等下个周期。
    if batch.get("status") != "completed":
        return
    if not batch.get("completed_at"):
        return

    # 二次 SSRF 校验（防 DB 被改）
    ok, reason = _is_safe_callback_url(callback_url)
    if not ok:
        await db.mark_callback_attempt(
            batch_id, success=False, error=f"unsafe_url:{reason}",
            max_attempts=1,  # 直接终态
        )
        return

    # 拉完成统计构造载荷
    snap = await db.get_batch_completion_status(batch_id)
    t = snap["tasks"]
    s = snap["screenshots"]
    total = t["total"] or 0
    success_rate = (t["done"] / total) if total > 0 else 0.0

    # 计算用时
    duration = None
    try:
        async with db._db.execute(
            "SELECT created_at FROM batches WHERE id=?", (batch_id,)
        ) as c:
            row2 = await c.fetchone()
        if row2 and row2[0] and batch.get("completed_at"):
            start = datetime.strptime(row2[0][:19], '%Y-%m-%d %H:%M:%S')
            stop = datetime.strptime(batch["completed_at"][:19], '%Y-%m-%d %H:%M:%S')
            duration = int((stop - start).total_seconds())
    except Exception:
        pass

    server_base = (_runtime_settings.get("server_public_base") or "").rstrip("/")
    if not server_base:
        # 没配置公开 URL，就用 callback_url 的同 scheme（仅作 hint，不影响发送）
        server_base = ""
    data_url = f"{server_base}/api/results?batch_id={batch_id}" if server_base else None
    export_url = f"{server_base}/api/export/{batch['name']}" if server_base else None

    attempts = (batch.get("callback_attempts") or 0) + 1
    completed_at_str = batch.get("completed_at") or ""
    event_id = (
        f"evt_{batch_id}_"
        f"{completed_at_str.replace(' ','T').replace(':','').replace('-','')}"
    ) if completed_at_str else f"evt_{batch_id}"

    payload = {
        "event": "batch.completed",
        "batch_id": batch_id,
        "batch_name": batch["name"],
        "external_id": batch.get("external_id"),
        "status": batch.get("status") or "completed",
        "stats": {
            "total": total,
            "done": t["done"] or 0,
            "failed": t["failed"] or 0,
            "success_rate": round(success_rate, 4),
            "duration_seconds": duration,
        },
        "screenshots": {
            "total": s["total"] or 0,
            "done": s["done"] or 0,
            "failed": s["failed"] or 0,
        },
        "completed_at": batch.get("completed_at"),
        "data_url": data_url,
        "export_url": export_url,
    }
    headers = {
        "X-Scraper-Event-Id": event_id,
        "X-Scraper-Delivery-Attempt": str(attempts),
        "User-Agent": "amazon-scraper-v3/callback",
        "Content-Type": "application/json",
    }

    try:
        resp = await client.post(callback_url, json=payload, headers=headers)
        if 200 <= resp.status_code < 300:
            await db.mark_callback_attempt(batch_id, success=True,
                                            max_attempts=_CALLBACK_MAX_ATTEMPTS)
            logger.info(f"📬 callback 发送成功 batch={batch_id} attempt={attempts}")
            return
        err = f"HTTP {resp.status_code}: {(resp.text or '')[:200]}"
    except asyncio.TimeoutError:
        err = "timeout"
    except Exception as e:
        err = f"{type(e).__name__}: {e}"[:300]

    # 失败处理：计算下次重试时间
    delay_idx = min(attempts - 1, len(_CALLBACK_RETRY_DELAYS) - 1)
    delay = _CALLBACK_RETRY_DELAYS[delay_idx] if attempts < _CALLBACK_MAX_ATTEMPTS else 0
    next_retry = (
        (datetime.utcnow() + timedelta(seconds=delay)).strftime('%Y-%m-%d %H:%M:%S')
        if attempts < _CALLBACK_MAX_ATTEMPTS else None
    )
    result = await db.mark_callback_attempt(
        batch_id, success=False, error=err,
        next_retry_at=next_retry, max_attempts=_CALLBACK_MAX_ATTEMPTS,
    )
    if result.get("final_status") == "failed":
        logger.warning(f"❌ callback 5 次失败终态 batch={batch_id} last_error={err}")
    else:
        logger.info(f"⏳ callback 失败 batch={batch_id} attempt={attempts}/{_CALLBACK_MAX_ATTEMPTS}，{delay}s 后重试")


async def _auto_scrape_scheduler():
    """定时自动采集调度（支持 interval_days 和文件指定 ASIN）"""
    while True:
        await asyncio.sleep(60)
        try:
            schedules = _runtime_settings.get("auto_scrape_schedules", [])
            now = _cn_now()  # 按中国时间判定「执行时间」，使 18:00 = 中国 18:00
            today = now.strftime("%Y-%m-%d")
            changed = False

            for sched in schedules:
                if not sched.get("enabled"):
                    continue
                sched_time = sched.get("time", "")
                if not sched_time:
                    continue
                try:
                    hour, minute = map(int, sched_time.split(":"))
                except ValueError:
                    continue

                # 检查时间点是否到达
                if now.hour < hour or (now.hour == hour and now.minute < minute):
                    continue

                # 检查间隔天数
                interval = sched.get("interval_days", 1)
                last_run = sched.get("last_run_date", "")
                if last_run:
                    try:
                        last_date = datetime.strptime(last_run, "%Y-%m-%d")
                        if (now - last_date).days < interval:
                            continue
                    except ValueError:
                        pass

                # 获取 ASIN 列表：优先从文件，否则全库
                source_file = sched.get("source_file", "")
                if source_file and os.path.isfile(source_file):
                    asins = _extract_asins_from_file(source_file)
                else:
                    asins = await db.get_all_asins()

                if not asins:
                    continue

                sched_name = sched.get("name", "task")
                batch_name = f"auto_{sched_name}_{now:%Y%m%d_%H%M}"
                zc = _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE)
                ns = sched.get("needs_screenshot", False)
                batch_id = await db.create_batch(batch_name, ns, is_auto=True)
                await db.create_tasks(batch_id, asins, zc, ns)
                sched["last_run_date"] = today
                changed = True
                logger.info(f"自动采集已调度: {batch_name}, {len(asins)} ASINs (间隔{interval}天)")

            if changed:
                _save_settings()
        except Exception as e:
            logger.error(f"自动采集调度异常: {e}")


# ==================== 全局并发协调 ====================

def _allocate_quotas():
    """根据 worker 健康度分配并发配额"""
    active = {wid: info for wid, info in _global_coordinator.items()
              if wid in _worker_registry}
    if not active:
        return

    max_conc = _runtime_settings.get("global_max_concurrency", config.GLOBAL_MAX_CONCURRENCY)
    max_qps = _runtime_settings.get("global_max_qps", config.GLOBAL_MAX_QPS)

    # 健康度加权分配
    scores = {}
    for wid, info in active.items():
        metrics = info.get("metrics", {})
        sr = metrics.get("success_rate", 1.0)
        br = metrics.get("block_rate", 0.0)
        scores[wid] = max(0.1, sr * (1 - br * 5))

    total_score = sum(scores.values())
    per_worker_max_conc = _runtime_settings.get("max_concurrency", config.MAX_CONCURRENCY)
    per_worker_max_qps = _runtime_settings.get("token_bucket_rate", config.TOKEN_BUCKET_RATE)

    for wid in active:
        weight = scores[wid] / total_score if total_score > 0 else 1.0 / len(active)
        allocated_conc = max(config.MIN_CONCURRENCY, int(max_conc * weight))
        allocated_qps = max(1.0, max_qps * weight)
        active[wid]["quota"] = {
            # 取 global 分配值和单 worker 上限的较小值
            "max_concurrency": min(allocated_conc, per_worker_max_conc),
            "max_qps": min(allocated_qps, per_worker_max_qps),
        }


# ==================== HTML 页面 ====================

@app.get("/", response_class=HTMLResponse)
async def page_dashboard(request: Request):
    progress = await db.get_progress()
    total = progress["done"] + progress["failed"]
    progress["completion_rate"] = round(progress["done"] / progress["total"] * 100, 1) if progress["total"] else 0
    progress["success_rate"] = round(progress["done"] / total * 100, 1) if total else 0
    batches = await db.get_batches()
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "progress": progress,
        "batches": batches,
        "active_workers": len([w for w in _worker_registry.values() if time.time() - w["last_seen"] < 60]),
    })


@app.get("/tasks", response_class=HTMLResponse)
async def page_tasks(request: Request):
    batches = await db.get_batches()
    return templates.TemplateResponse("tasks.html", {
        "request": request,
        "batches": batches,
        "default_zip_code": _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE),
    })


@app.get("/results", response_class=HTMLResponse)
async def page_results(request: Request):
    return templates.TemplateResponse("results.html", {"request": request})


@app.get("/workers", response_class=HTMLResponse)
async def page_workers(request: Request):
    return templates.TemplateResponse("workers.html", {"request": request})


@app.get("/settings", response_class=HTMLResponse)
async def page_settings(request: Request):
    return templates.TemplateResponse("settings.html", {
        "request": request,
        "settings": _runtime_settings,
        "config": {
            "port": config.SERVER_PORT,
            "timeout": config.REQUEST_TIMEOUT,
            "db_path": config.DB_PATH,
        },
    })


# ==================== API: 批次和任务 ====================

MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50MB 上限，防止 2GB 内存 VPS OOM


# ==================== 完成通知 / Callback 基础设施 ====================

# 内存中需要"检查是否完成"的 batch_id 集合：
# - 由 accept_results_batch 在写入结果后 add(batch_id) （O(1)、不阻塞）
# - 由 _completion_watcher 协程消费
_completion_check_set: Set[int] = set()
# 变体自动展开：记录每个批次已展开的轮次。正常 2 轮收敛（轮1=上传，轮2=变体）；
# 若第 2 次展开仍新增（=采集第 3 轮）则打 WARNING（家族列表不一致信号），但靠
# UNIQUE(batch_id,asin) 去重仍保证收敛、不死循环。批次完成后清理。
_expand_rounds: Dict[int, int] = {}
# 待发送的 callback 内存队列：
# - _completion_watcher 把刚标记 completed 的 batch_id 入队
# - _timeout_task_loop 兜底从 DB 扫描入队（处理重启 / 实时漏掉的）
# - _callback_dispatcher 串行消费 + HTTP 发送
_callback_send_queue: "asyncio.Queue[int]" = None  # 在 lifespan 启动时实例化

# 回调重试退避（秒）
_CALLBACK_RETRY_DELAYS = [30, 300, 1800, 7200]   # 第 1/2/3/4 次失败后的等待
_CALLBACK_MAX_ATTEMPTS = 5
_CALLBACK_HTTP_TIMEOUT = 10                       # 单次 HTTP 超时
_COMPLETION_WATCHER_INTERVAL = 2.0                # 每 2 秒批量处理一次

# SSRF 防护：拒绝指向内网的 URL
_SSRF_BLOCKED_HOSTS = {"localhost", "ip6-localhost", "ip6-loopback"}


def _is_safe_callback_url(url: str) -> tuple[bool, str]:
    """判定 callback URL 是否安全可用于公网回调。

    返回 (ok, reason)。公网场景下：
    - 只允许 http/https
    - 拒绝 localhost 字面量
    - 拒绝字面量私网 / 回环 / 链路本地 IP（10/8、172.16-31、192.168/16、127/8、169.254/16、::1 等）
    - 域名一律放行（公网假设；不做 DNS 解析以避免 TOCTOU）
    """
    if not url or not isinstance(url, str):
        return False, "empty"
    try:
        u = urlparse(url.strip())
    except Exception as e:
        return False, f"invalid_url:{e}"
    if u.scheme not in ("http", "https"):
        return False, f"bad_scheme:{u.scheme}"
    host = (u.hostname or "").lower()
    if not host:
        return False, "missing_host"
    if host in _SSRF_BLOCKED_HOSTS:
        return False, "blocked_host"
    # 尝试作为 IP 字面量解析
    try:
        ip = ipaddress.ip_address(host)
        if ip.is_loopback or ip.is_private or ip.is_link_local or ip.is_reserved or ip.is_multicast:
            return False, f"blocked_ip:{host}"
    except ValueError:
        pass  # 是域名
    return True, ""


_ASIN_RE = re.compile(r'^B[0-9A-Z]{9}$')
# 美国邮编：5 位数字（兼容 ZIP+4，前 5 位）
_US_ZIP_RE = re.compile(r'^\d{5}$')


def _normalize_asin(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip().upper()
    return s if _ASIN_RE.match(s) else None


def _normalize_zip(val) -> Optional[str]:
    """规范化邮编。返回 5 位数字字符串，无效则 None。
    支持去掉前导/尾随空白、ZIP+4（取前 5 位）、Excel 数字单元格（如 10001 不会带前导 0 的损失）。
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Excel 数字 90001.0 → "90001"
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    # 前 5 位（兼容 "10001-1234"）
    head = s.split("-", 1)[0].strip()
    # 数字邮编位数补 0（Excel 把 "01234" 存为 1234）
    if head.isdigit() and len(head) <= 5:
        head = head.zfill(5)
    return head if _US_ZIP_RE.match(head) else None


@app.post("/api/upload")
async def api_upload(request: Request,
                     file: UploadFile = File(...),
                     batch_name: str = Form(None),
                     zip_code: str = Form(None),
                     needs_screenshot: bool = Form(False),
                     callback_url: str = Form(None),
                     external_id: str = Form(None),
                     expand_variants: bool = Form(False)):
    """上传 ASIN 文件创建批次。

    支持 xlsx / csv / txt 三种格式：
    - **xlsx / csv**：A 列 = ASIN，B 列（可选）= 该 ASIN 单独的采集邮编（5 位数字）
      B 列为空 → 用本次上传指定的 batch zip（或服务端默认）
    - **txt**：每行一个 ASIN（无邮编列）

    可选 callback：
    - `callback_url`：批次完成时（含截图）POST 到此 URL 通知调用方
    - `external_id`：调用方自己的批次 ID，原样回传，便于追踪
    """
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"文件过大：{len(content)//1024//1024}MB，上限 {MAX_UPLOAD_BYTES//1024//1024}MB")
    filename = file.filename or ""

    asins: List[str] = []                # 顺序、可重复（去重在后面）
    per_asin_zip: Dict[str, str] = {}    # 仅记录 B 列指定的邮编
    invalid_zip_count = 0

    def add_pair(asin_val, zip_val):
        nonlocal invalid_zip_count
        asin = _normalize_asin(asin_val)
        if not asin:
            return
        asins.append(asin)
        if zip_val is not None and str(zip_val).strip():
            zip_norm = _normalize_zip(zip_val)
            if zip_norm:
                # 同一 ASIN 在多行重复时，以第一次出现的非空 zip 为准
                per_asin_zip.setdefault(asin, zip_norm)
            else:
                invalid_zip_count += 1

    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row:
                    continue
                a = row[0] if len(row) > 0 else None
                b = row[1] if len(row) > 1 else None
                # 兼容旧上传：当 A 列不是 ASIN 时，扫描该行所有列找 ASIN（不带 zip）
                if _normalize_asin(a):
                    add_pair(a, b)
                else:
                    for cell in row:
                        if _normalize_asin(cell):
                            add_pair(cell, None)
        finally:
            wb.close()
    elif filename.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if not row:
                continue
            a = row[0] if len(row) > 0 else None
            b = row[1] if len(row) > 1 else None
            if _normalize_asin(a):
                add_pair(a, b)
            else:
                for cell in row:
                    if _normalize_asin(cell):
                        add_pair(cell, None)
    else:
        text = content.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            add_pair(line, None)

    if not asins:
        raise HTTPException(400, "未找到有效 ASIN")

    # 去重（保持顺序）
    seen = set()
    unique_asins = []
    for a in asins:
        if a not in seen:
            unique_asins.append(a)
            seen.add(a)

    if not batch_name:
        batch_name = f"batch_{_cn_now().strftime('%Y%m%d_%H%M%S')}"

    # callback_url 校验（防 SSRF + 格式）
    cb_url = (callback_url or "").strip() or None
    if cb_url:
        ok, reason = _is_safe_callback_url(cb_url)
        if not ok:
            raise HTTPException(400, f"非法 callback_url（{reason}）。仅接受 http(s)://公网域名/IP")

    ext_id = (external_id or "").strip()[:120] or None  # 上限 120 字符防滥用

    zc = zip_code or _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE)
    batch_id = await db.create_batch(
        batch_name, needs_screenshot,
        external_id=ext_id, callback_url=cb_url,
        expand_variants=expand_variants,
    )
    inserted = await db.create_tasks(
        batch_id, unique_asins, zc, needs_screenshot,
        per_asin_zip=per_asin_zip,
    )

    # 构造 status URL（调用方可以轮询）
    base = str(request.base_url).rstrip("/")
    return {
        "batch_id": batch_id,
        "batch_name": batch_name,
        "external_id": ext_id,
        "total_asins": len(unique_asins),
        "inserted": inserted,
        "per_asin_zip_count": len(per_asin_zip),
        "invalid_zip_rows": invalid_zip_count,
        "callback_url": cb_url,
        "status_url": f"{base}/api/batches/{batch_name}/status",
    }


# ==================== F-009: 卖家店铺采集 ====================

# Amazon 三方卖家 ID 通常是 13-14 位 A 开头的字母数字串（如 A2L77EE7U53NWQ）。
# 同时支持从 URL 中提取 me=... / seller=... 参数。
_SELLER_URL_RE = re.compile(r'(?:[?&](?:me|seller)=)([A-Z0-9]{10,16})', re.IGNORECASE)
_BARE_SELLER_RE = re.compile(r'^A[A-Z0-9]{12,14}$')


def _extract_sellers_from_text(text: str) -> List[str]:
    """从一段文本中提取所有 seller_id（去重，保持出现顺序）。

    支持 3 种形式：
      1. 完整 URL: https://www.amazon.com/s?me=A2L77EE7U53NWQ
      2. URL 片段: ?me=A2L77... 或 ?seller=A2L77...
      3. 裸 ID: A2L77EE7U53NWQ
    """
    seen = set()
    out = []
    for line in text.splitlines():
        candidates = []
        for m in _SELLER_URL_RE.finditer(line):
            candidates.append(m.group(1).upper())
        for tok in re.split(r'[\s,;\t]+', line):
            tok = tok.strip().upper()
            if _BARE_SELLER_RE.match(tok):
                candidates.append(tok)
        for sid in candidates:
            if sid not in seen:
                seen.add(sid)
                out.append(sid)
    return out


@app.post("/api/upload-sellers")
async def api_upload_sellers(request: Request,
                              file: UploadFile = File(...),
                              batch_name: str = Form(None),
                              discover_mode: str = Form("with_detail"),
                              zip_code: str = Form(None),
                              needs_screenshot: bool = Form(False)):
    """上传卖家 ID/URL 文件，创建一个 seller_discovery 批次。

    discover_mode:
      - 'discover_only': 仅翻页发现 ASIN，写入 seller_discoveries 表
      - 'with_detail':   发现后自动衍生 ASIN 详情任务进入主采集队列
    """
    if discover_mode not in ("discover_only", "with_detail"):
        raise HTTPException(400, f"非法 discover_mode: {discover_mode}")

    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"文件过大：{len(content)//1024//1024}MB，上限 {MAX_UPLOAD_BYTES//1024//1024}MB")
    filename = (file.filename or "").lower()

    seller_ids: List[str] = []
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        try:
            ws = wb.active
            buf = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                for cell in row:
                    if cell:
                        buf.append(str(cell))
            seller_ids = _extract_sellers_from_text("\n".join(buf))
        finally:
            wb.close()
    elif filename.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        seller_ids = _extract_sellers_from_text(text)
    else:
        text = content.decode("utf-8", errors="ignore")
        seller_ids = _extract_sellers_from_text(text)

    if not seller_ids:
        raise HTTPException(400, "未识别到任何 seller ID（支持裸 ID、含 me=/seller= 的 URL）")

    if not batch_name:
        batch_name = f"sellers_{_cn_now().strftime('%Y%m%d_%H%M%S')}"

    zc = zip_code or _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE)
    batch_id, inserted = await db.create_seller_batch(
        name=batch_name,
        seller_ids=seller_ids,
        discover_mode=discover_mode,
        zip_code=zc,
        needs_screenshot=needs_screenshot,
    )
    if not batch_id:
        raise HTTPException(500, "创建卖家批次失败")

    return {
        "batch_id": batch_id,
        "batch_name": batch_name,
        "discover_mode": discover_mode,
        "total_sellers": len(seller_ids),
        "inserted_tasks": inserted,
    }


@app.get("/api/seller-batches/{batch_id}/progress")
async def api_seller_batch_progress(batch_id: int):
    """seller_discovery 批次专属进度端点：discover + detail + 已发现 ASIN 数。"""
    return await db.get_seller_batch_progress(batch_id)


@app.get("/api/seller-batches/{batch_id}/discoveries")
async def api_seller_discoveries(batch_id: int,
                                  seller_id: Optional[str] = None,
                                  limit: int = 200,
                                  offset: int = 0):
    """列出某批次发现的 ASIN（可按 seller_id 过滤）。"""
    limit = max(1, min(limit, 1000))
    offset = max(0, offset)
    where = ["batch_id = ?"]
    params: List[Any] = [batch_id]
    if seller_id:
        where.append("seller_id = ?")
        params.append(seller_id.strip().upper())
    sql = (
        "SELECT seller_id, asin, list_title, list_price, list_image, discovered_at "
        f"FROM seller_discoveries WHERE {' AND '.join(where)} "
        "ORDER BY discovered_at DESC, asin ASC LIMIT ? OFFSET ?"
    )
    params.extend([limit, offset])
    rows = []
    async with db._db.execute(sql, params) as c:
        async for r in c:
            rows.append(dict(r))
    return {"items": rows, "limit": limit, "offset": offset}


@app.get("/api/batches")
async def api_batches():
    batches = await db.get_batches()
    return {"batches": batches}


@app.get("/api/progress")
async def api_progress(batch_id: int = None):
    return await db.get_progress(batch_id)


@app.get("/api/batches/{batch_name}/screenshots/progress")
async def api_screenshot_progress(batch_name: str):
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    return await db.get_screenshot_progress(batch["id"])


@app.get("/api/batches/{batch_name}/status")
async def api_batch_status(batch_name: str):
    """轮询批次完成状态（兼容旧调用方 + 给纯轮询模式用）。

    返回结构：
        {
          "batch_name": ...,
          "batch_id": ...,
          "external_id": ...,
          "status": "running"|"completed"|"failed",
          "stats": {total, done, failed, success_rate, duration_seconds},
          "screenshots": {total, done, failed},
          "completed_at": null|"2026-...",
          "callback": {url, status, attempts, last_error, sent_at}
        }
    """
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    snap = await db.get_batch_completion_status(batch["id"])
    t = snap["tasks"]
    s = snap["screenshots"]
    total = t["total"] or 0
    success_rate = (t["done"] / total) if total > 0 else 0.0

    # 持续时长（创建到完成；未完成则到当前时间）
    duration = None
    try:
        created_at = batch.get("created_at")
        end_at = batch.get("completed_at") or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        if created_at:
            start = datetime.strptime(created_at[:19], '%Y-%m-%d %H:%M:%S')
            stop = datetime.strptime(end_at[:19], '%Y-%m-%d %H:%M:%S')
            duration = int((stop - start).total_seconds())
    except Exception:
        pass

    return {
        "batch_name": batch_name,
        "batch_id": batch["id"],
        "external_id": batch.get("external_id"),
        "status": batch.get("status") or "running",
        "stats": {
            "total": total,
            "done": t["done"] or 0,
            "failed": t["failed"] or 0,
            "open": t["open"] or 0,
            "success_rate": round(success_rate, 4),
            "duration_seconds": duration,
        },
        "screenshots": {
            "total": s["total"] or 0,
            "done": s["done"] or 0,
            "failed": s["failed"] or 0,
            "open": s["open"] or 0,
        },
        "completed_at": batch.get("completed_at"),
        "callback": {
            "url": batch.get("callback_url"),
            "status": batch.get("callback_status"),
            "attempts": batch.get("callback_attempts") or 0,
            "last_error": batch.get("callback_last_error"),
            "sent_at": batch.get("callback_sent_at"),
            "next_retry_at": batch.get("callback_next_retry_at"),
        },
    }


@app.post("/api/batches/{batch_name}/callback/retry")
async def api_batch_callback_retry(batch_name: str):
    """运维手动触发：重置该批次的 callback 状态，立即重新发送。
    可用于：webhook 失败 5 次后；或调用方端点恢复后想重新接收一次通知。
    """
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    if not batch.get("callback_url"):
        raise HTTPException(400, "该批次没有配置 callback_url")
    changed = await db.reset_callback_for_retry(batch["id"])
    # 入队让 dispatcher 立刻处理
    if _callback_send_queue is not None:
        try:
            _callback_send_queue.put_nowait(batch["id"])
        except Exception:
            pass
    return {"ok": changed, "batch_id": batch["id"]}


@app.post("/api/batches/{batch_id}/prioritize")
async def api_prioritize(batch_id: int):
    await db.prioritize_batch(batch_id)
    return {"ok": True}


@app.post("/api/batches/{batch_name}/retry")
async def api_retry_batch(batch_name: str, force: bool = False):
    """重试失败任务

    默认跳过 NO_AUTO_RETRY_ERROR_TYPES（如 variant_offset），因为这些类型
    已经在 layer 2 给过有限次重试机会，仍失败说明是 Amazon 侧稳定问题。
    如确需强制重试，可加 ?force=true（前端按住 Shift 点击）。
    返回 retried/skipped_no_retry 数量，前端可展示。
    """
    from common.database import NO_AUTO_RETRY_ERROR_TYPES
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    batch_id = batch["id"]

    no_retry_list = sorted(NO_AUTO_RETRY_ERROR_TYPES)
    no_retry_placeholders = ",".join("?" * len(no_retry_list))

    async with db._write_lock:
        await db._db.execute("BEGIN")
        # 先统计：本次将跳过的"不可重试"任务数（供前端展示）
        skipped = 0
        if not force and no_retry_list:
            async with db._db.execute(
                f"SELECT COUNT(*) FROM tasks WHERE batch_id=? AND status='failed' "
                f"AND COALESCE(error_type, '') IN ({no_retry_placeholders})",
                [batch_id] + no_retry_list
            ) as c:
                row = await c.fetchone()
                skipped = row[0] if row else 0

        # 实际重试 SQL：默认排除 NO_RETRY；force=true 时全部重试
        if force or not no_retry_list:
            cursor = await db._db.execute(
                "UPDATE tasks SET status='pending', retry_count=0, worker_id=NULL "
                "WHERE batch_id=? AND status='failed'",
                (batch_id,)
            )
        else:
            cursor = await db._db.execute(
                f"UPDATE tasks SET status='pending', retry_count=0, worker_id=NULL "
                f"WHERE batch_id=? AND status='failed' "
                f"AND COALESCE(error_type, '') NOT IN ({no_retry_placeholders})",
                [batch_id] + no_retry_list
            )
        retried = cursor.rowcount
        await db._db.execute("COMMIT")
    return {
        "ok": True,
        "retried": retried,
        "skipped_no_retry": skipped,
        "no_retry_types": no_retry_list,
        "forced": force,
    }


@app.delete("/api/batches/{batch_name}")
async def api_delete_batch(batch_name: str):
    """删除批次及其任务"""
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    batch_id = batch["id"]
    # 先收集截图物理文件路径
    screenshot_files = []
    async with db._db.execute(
        "SELECT file_path FROM screenshots WHERE batch_id=? AND file_path IS NOT NULL", (batch_id,)
    ) as c:
        screenshot_files = [row["file_path"] for row in await c.fetchall()]

    async with db._write_lock:
        await db._db.execute("BEGIN")
        await db._db.execute("DELETE FROM tasks WHERE batch_id=?", (batch_id,))
        await db._db.execute("DELETE FROM batch_asins WHERE batch_id=?", (batch_id,))
        await db._db.execute("DELETE FROM screenshots WHERE batch_id=?", (batch_id,))
        await db._db.execute("DELETE FROM asin_changes WHERE batch_id=?", (batch_id,))
        await db._db.execute("DELETE FROM batches WHERE id=?", (batch_id,))
        await db._db.execute("COMMIT")

    # 删除物理截图文件
    _remove_screenshot_files(screenshot_files)
    return {"ok": True}


@app.post("/api/batches/delete-bulk")
async def api_delete_batches_bulk(request: Request):
    """批量删除多个批次（按 batch_id）及其全部关联数据 + 截图文件。
    入参 JSON：{"batch_ids": [1,2,3]}。一次事务删除，原子性。"""
    body = await request.json()
    raw = body.get("batch_ids", [])
    if not isinstance(raw, list):
        raise HTTPException(400, "batch_ids 必须是数组")
    # 仅接受整数 id，去重，上限保护（防超大 IN 子句）
    seen = set()
    batch_ids = []
    for x in raw:
        try:
            i = int(x)
        except (ValueError, TypeError):
            continue
        if i not in seen:
            seen.add(i)
            batch_ids.append(i)
    batch_ids = batch_ids[:500]
    if not batch_ids:
        raise HTTPException(400, "batch_ids 为空或无效")

    ph = ",".join("?" * len(batch_ids))
    # 先收集截图物理文件路径（只读连接，不占写锁）
    async with db.read() as rc, rc.execute(
        f"SELECT file_path FROM screenshots WHERE batch_id IN ({ph}) AND file_path IS NOT NULL",
        batch_ids
    ) as c:
        screenshot_files = [row["file_path"] for row in await c.fetchall()]

    async with db._write_lock:
        await db._db.execute("BEGIN")
        try:
            for tbl in ("tasks", "batch_asins", "screenshots", "asin_changes"):
                await db._db.execute(f"DELETE FROM {tbl} WHERE batch_id IN ({ph})", batch_ids)
            await db._db.execute(f"DELETE FROM batches WHERE id IN ({ph})", batch_ids)
            await db._db.execute("COMMIT")
        except Exception:
            try:
                await db._db.execute("ROLLBACK")
            except Exception:
                pass
            raise

    _remove_screenshot_files(screenshot_files)
    logger.info(f"批量删除批次: {len(batch_ids)} 个 (ids={batch_ids[:20]}{'...' if len(batch_ids) > 20 else ''})")
    return {"ok": True, "deleted": len(batch_ids)}


@app.get("/api/batches/{batch_name}/errors")
async def api_batch_errors(batch_name: str):
    """获取批次错误详情"""
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    batch_id = batch["id"]
    async with db.read() as rc:
        async with rc.execute(
            "SELECT error_type, COUNT(*) as cnt FROM tasks "
            "WHERE batch_id=? AND status='failed' "
            "GROUP BY error_type ORDER BY cnt DESC",
            (batch_id,)
        ) as c:
            error_summary = [dict(r) for r in await c.fetchall()]
        async with rc.execute(
            "SELECT asin, error_type, error_detail, retry_count, worker_id, updated_at "
            "FROM tasks WHERE batch_id=? AND status='failed' "
            "ORDER BY updated_at DESC LIMIT 200",
            (batch_id,)
        ) as c:
            failed_tasks = [dict(r) for r in await c.fetchall()]
    return {"error_summary": error_summary, "failed_tasks": failed_tasks}


@app.get("/api/coordinator")
async def api_coordinator():
    """全局并发协调器状态"""
    max_conc = _runtime_settings.get("global_max_concurrency", config.GLOBAL_MAX_CONCURRENCY)
    max_qps = _runtime_settings.get("global_max_qps", config.GLOBAL_MAX_QPS)
    allocated_conc = sum(info.get("quota", {}).get("max_concurrency", 0)
                        for info in _global_coordinator.values())
    allocated_qps = sum(info.get("quota", {}).get("max_qps", 0)
                        for info in _global_coordinator.values())
    return {
        "max_concurrency": max_conc,
        "allocated_concurrency": allocated_conc,
        "max_qps": max_qps,
        "allocated_qps": allocated_qps,
        "active_workers": len(_global_coordinator),
        "global_block_until": 0,
        "global_block_count": 0,
    }


@app.delete("/api/workers")
async def api_delete_all_offline():
    """清除所有离线 worker"""
    cutoff = time.time() - 60
    offline = [wid for wid, w in _worker_registry.items() if w["last_seen"] < cutoff]
    for wid in offline:
        del _worker_registry[wid]
        _global_coordinator.pop(wid, None)
    return {"ok": True, "removed": len(offline)}


@app.post("/api/workers/{worker_id}/restart")
async def api_restart_worker(worker_id: str):
    """标记 worker 软重启（下次 sync 时生效）"""
    _worker_restart_flags[worker_id] = True
    return {"ok": True, "message": f"重启指令已下发，{worker_id} 将在 30 秒内重启"}


@app.post("/api/settings/reset")
async def api_reset_settings():
    """恢复默认设置"""
    global _runtime_settings, _settings_version
    _runtime_settings = _default_settings()
    _settings_version += 1
    _save_settings()
    return {"ok": True, "settings": _runtime_settings}


# ==================== API: Worker 任务拉取和提交 ====================

@app.get("/api/tasks/pull")
async def api_pull_tasks(request: Request,
                         worker_id: str = Query(...),
                         count: int = Query(10),
                         needs_screenshot: Optional[bool] = Query(None),
                         enable_screenshot: Optional[bool] = Query(None),
                         prefer_zip: Optional[str] = Query(None)):
    ip = request.client.host if request.client else None
    # enable_screenshot 来自 worker 的 query param，更新 registry
    _register_worker(worker_id, enable_screenshot=enable_screenshot, ip=ip)

    # ns=None: 不限制; ns=False: 只拉不需要截图的任务
    # enable_screenshot 表示 worker 是否有截图能力：
    #   True  → 拉所有任务 (ns=None)
    #   False → 只拉不需要截图的 (ns=False)
    ns = None
    if needs_screenshot is not None:
        ns = needs_screenshot
    elif enable_screenshot is not None:
        if not enable_screenshot:
            ns = False
        # enable_screenshot=True → ns=None (拉所有任务)
    elif worker_id in _worker_registry:
        if not _worker_registry[worker_id].get("enable_screenshot", True):
            ns = False

    # prefer_zip 校验（防注入）：只接受 5 位数字
    pz = None
    if prefer_zip and re.match(r'^\d{5}$', prefer_zip):
        pz = prefer_zip

    tasks = await db.pull_tasks(worker_id, count, ns, prefer_zip=pz)
    if worker_id in _worker_registry:
        _worker_registry[worker_id]["tasks_pulled"] += len(tasks)
    return {"tasks": tasks}


@app.post("/api/tasks/release")
async def api_release_tasks(request: Request):
    body = await request.json()
    worker_id = body.get("worker_id", "")
    tasks = body.get("tasks", [])
    # 兼容旧格式 {"task_ids": [1,2,3]}（无 lease 校验，直接释放）
    if not tasks and "task_ids" in body:
        task_ids = body["task_ids"]
        if task_ids:
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            placeholders = ",".join("?" * len(task_ids))
            async with db._write_lock:
                await db._db.execute("BEGIN")
                cursor = await db._db.execute(
                    f"UPDATE tasks SET status='pending', worker_id=NULL, "
                    f"lease_epoch=lease_epoch+1, updated_at=? "
                    f"WHERE id IN ({placeholders}) AND status='processing'",
                    [now] + task_ids
                )
                await db._db.execute("COMMIT")
            return {"ok": True, "released": cursor.rowcount}
        return {"ok": True, "released": 0}
    released = await db.release_tasks(worker_id, tasks)
    return {"ok": True, "released": released}


@app.post("/api/tasks/result")
async def api_submit_result(request: Request):
    """提交单个结果（lease 校验 → 原子写入）"""
    data = await request.json()
    task_id = data.pop("task_id", None)
    batch_id = data.pop("batch_id", None)
    worker_id = data.get("worker_id", "")
    lease_epoch = data.pop("lease_epoch", 0)

    if task_id:
        if data.get("success", True):
            result = await db.accept_success_result(task_id, worker_id, lease_epoch, data, batch_id)
        else:
            result = await db.accept_failed_result(
                task_id, worker_id, lease_epoch,
                data.get("error_type", ""), data.get("error_detail", ""))
        if worker_id in _worker_registry and result.get("accepted"):
            _worker_registry[worker_id]["results_submitted"] += 1
        return {"ok": result.get("accepted", False), "stale": result.get("stale", False)}
    else:
        # 无 task_id 的直接写入（兼容）
        saved = await db.save_result(data, batch_id)
        return {"ok": saved, "stale": False}


@app.post("/api/tasks/result/batch")
async def api_submit_batch(request: Request):
    """批量提交结果（单事务，减少锁争用）"""
    body = await request.json()
    results = body.get("results", [])

    # 构建 batch items
    batch_items = []
    worker_id_set = set()
    for item in results:
        task_id = item.pop("task_id", None)
        batch_id = item.pop("batch_id", None)
        worker_id = item.get("worker_id", "")
        lease_epoch = item.pop("lease_epoch", 0)
        is_success = item.pop("success", True)
        worker_id_set.add(worker_id)
        batch_items.append({
            "task_id": task_id,
            "worker_id": worker_id,
            "lease_epoch": lease_epoch,
            "batch_id": batch_id,
            "data": item,
            "success": is_success,
        })

    result = await db.accept_results_batch(batch_items)

    # results_submitted 只计 accepted
    for wid in worker_id_set:
        if wid in _worker_registry and result["accepted"] > 0:
            _worker_registry[wid]["results_submitted"] += result["accepted"]

    # 防御性入队：本次写入涉及的 batch_id 标记为"需要检查是否完成"。
    # set.add 是 O(1) 内存操作，不会影响 worker 提交响应。
    # 任何异常都吞掉只 log，绝不让通知机制污染采集主路径。
    try:
        touched = {item["batch_id"] for item in batch_items if item.get("batch_id")}
        for bid in touched:
            _completion_check_set.add(bid)
    except Exception as e:
        logger.warning(f"完成检测入队异常（不影响采集）: {e}")

    return {**result, "total": len(results)}


@app.post("/api/tasks/seller-result")
async def api_submit_seller_result(request: Request):
    """接收 worker 的 discover_seller 任务结果（F-009）。

    Payload: {task_id, batch_id, worker_id, lease_epoch, seller_id, items, meta}
    """
    body = await request.json()
    task_id = body.get("task_id")
    batch_id = body.get("batch_id")
    worker_id = body.get("worker_id", "")
    lease_epoch = body.get("lease_epoch", 0)
    seller_id = (body.get("seller_id") or "").strip().upper()
    items = body.get("items") or []
    meta = body.get("meta") or {}

    if not task_id or not seller_id:
        raise HTTPException(400, "task_id 和 seller_id 必填")

    result = await db.accept_seller_discovery_result(
        task_id=task_id,
        worker_id=worker_id,
        lease_epoch=lease_epoch,
        batch_id=batch_id,
        seller_id=seller_id,
        items=items,
        meta=meta,
    )
    if worker_id in _worker_registry and result.get("accepted"):
        _worker_registry[worker_id]["results_submitted"] += 1
    return result


@app.post("/api/tasks/screenshot")
async def api_upload_screenshot(request: Request,
                                asin: str = Form(...),
                                batch_name: str = Form(...),
                                worker_id: str = Form(""),
                                file: UploadFile = File(...)):
    """接收截图上传（检查 worker 存活状态）"""
    # 拒绝已死 worker 的截图上传
    if worker_id and worker_id not in _worker_registry:
        raise HTTPException(409, f"Worker {worker_id} 已离线，截图被丢弃")
    if worker_id and worker_id in _worker_registry:
        if time.time() - _worker_registry[worker_id]["last_seen"] > 120:
            raise HTTPException(409, f"Worker {worker_id} 心跳超时，截图被丢弃")

    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(400, f"批次不存在: {batch_name}")

    batch_id = batch["id"]
    save_dir = os.path.join(config.SCREENSHOT_DIR, batch_name)
    os.makedirs(save_dir, exist_ok=True)

    filename = f"{asin}.png"
    filepath = os.path.join(save_dir, filename)
    content = await file.read()
    # 单张截图上限 10MB，防止恶意/损坏文件耗尽磁盘
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(413, f"截图过大：{len(content)//1024//1024}MB，上限 10MB")
    with open(filepath, "wb") as f:
        f.write(content)

    rel_path = f"/static/screenshots/{batch_name}/{filename}"
    updated = await db.update_screenshot_status(asin, batch_id, "done", file_path=rel_path)
    if not updated:
        try:
            os.remove(filepath)
        except OSError:
            pass
        raise HTTPException(409, f"截图状态不存在: {asin}@{batch_name}")
    return {"ok": True, "path": rel_path}


@app.post("/api/tasks/screenshot/fail")
async def api_screenshot_fail(request: Request):
    """截图渲染失败上报（触发重试或标记永久失败）"""
    body = await request.json()
    asin = body.get("asin", "")
    batch_name = body.get("batch_name", "")
    error = body.get("error", "unknown")
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(400, f"批次不存在: {batch_name}")
    updated = await db.update_screenshot_status(asin, batch["id"], "failed", error=error)
    if not updated:
        raise HTTPException(409, f"截图状态不存在: {asin}@{batch_name}")
    return {"ok": True}


# ==================== API: Worker 同步 ====================

@app.post("/api/worker/sync")
async def api_worker_sync(request: Request):
    """Worker 心跳 + 指标上报 + 设置/配额下发"""
    body = await request.json()
    worker_id = body.get("worker_id", "")
    ip = request.client.host if request.client else None

    _register_worker(worker_id, body.get("enable_screenshot"), ip)

    # 更新指标
    metrics = body.get("metrics", {})
    if worker_id not in _global_coordinator:
        _global_coordinator[worker_id] = {"metrics": {}, "quota": {}}
    _global_coordinator[worker_id]["metrics"] = metrics
    _allocate_quotas()

    quota = _global_coordinator.get(worker_id, {}).get("quota", {})

    # 检查重启标记
    restart = _worker_restart_flags.pop(worker_id, False)

    return {
        "settings": _runtime_settings,
        "settings_version": _settings_version,
        "quota": quota,
        "restart": restart,
    }


@app.get("/api/workers")
async def api_workers():
    now = time.time()
    workers = []
    for wid, w in _worker_registry.items():
        coord = _global_coordinator.get(wid, {})
        metrics = coord.get("metrics", {})
        quota = coord.get("quota", {})
        uptime = now - w.get("first_seen", now)
        workers.append({
            **w,
            "online": (now - w["last_seen"]) < 60,
            "last_seen_ago": int(now - w["last_seen"]),
            "uptime": int(uptime),
            "success_rate": metrics.get("success_rate"),
            "block_rate": metrics.get("block_rate"),
            "latency_p50": metrics.get("latency_p50"),
            "inflight": metrics.get("inflight"),
            "task_queue_size": metrics.get("task_queue_size"),
            "result_queue_size": metrics.get("result_queue_size"),
            "accepted": metrics.get("accepted", 0),
            "stale": metrics.get("stale", 0),
            "quota_concurrency": quota.get("max_concurrency"),
            "quota_qps": quota.get("max_qps"),
        })
    return {"workers": workers}


@app.delete("/api/workers/{worker_id}")
async def api_delete_worker(worker_id: str):
    _worker_registry.pop(worker_id, None)
    _global_coordinator.pop(worker_id, None)
    return {"ok": True}


# ==================== API: 结果查询 ====================

@app.get("/api/results")
async def api_results(batch_id: int = None,
                      cursor: int = None,
                      limit: int = Query(50, le=200),
                      search: str = None,
                      change_filter: str = "all",
                      direction: str = "next"):
    result = await db.get_results(
        batch_id=batch_id,
        cursor_id=cursor,
        limit=limit,
        search=search,
        change_filter=change_filter,
        direction=direction,
    )
    return result


@app.get("/api/results/{asin}")
async def api_result_detail(asin: str):
    data = await db.get_result_by_asin(asin)
    if not data:
        raise HTTPException(404, f"ASIN {asin} 不存在")
    changes = await db.get_asin_changes(asin)
    return {"data": data, "changes": changes}


@app.get("/api/changes/stats")
async def api_change_stats(batch_id: int = None):
    return await db.get_change_stats(batch_id)


# ==================== API: 设置 ====================

@app.get("/api/settings")
async def api_get_settings():
    return {"settings": _runtime_settings, "version": _settings_version}


@app.put("/api/settings")
async def api_update_settings(request: Request):
    global _settings_version
    body = await request.json()

    for key, value in body.items():
        if key in _runtime_settings:
            _runtime_settings[key] = value

    _settings_version += 1
    _save_settings()
    return {"ok": True, "version": _settings_version, "settings": _runtime_settings}


# ==================== API: 导出 ====================

from common.models import EXPORTABLE_FIELDS
from common.database import _parse_price_float


def _parse_selected_fields(fields_param: str = None):
    """解析并校验字段选择，返回 None 表示全选"""
    if not fields_param:
        return None
    selected = [f for f in fields_param.split(",") if f in EXPORTABLE_FIELDS]
    return selected if selected else None


def _get_export_headers(selected_fields=None):
    """构建导出表头和字段键"""
    if selected_fields is None:
        selected_fields = list(EXPORTABLE_FIELDS)

    include_total = "total_price" in selected_fields
    field_keys = [f for f in selected_fields if f != "total_price"]
    headers = [config.HEADER_MAP.get(f, f) for f in field_keys]

    if include_total:
        shipping_h = config.HEADER_MAP.get("buybox_shipping", "buybox_shipping")
        idx = headers.index(shipping_h) + 1 if shipping_h in headers else len(headers)
        headers.insert(idx, config.HEADER_MAP.get("total_price", "总价"))

    return headers, field_keys, include_total


def _prepare_row(item: dict, field_keys: list, headers: list, include_total: bool):
    """构建单行导出数据"""
    row = [str(item.get(f, "") or "") for f in field_keys]
    if include_total:
        bp = _parse_price_float(item.get("buybox_price", ""))
        bs_str = str(item.get("buybox_shipping", ""))
        if bp is not None:
            bs = 0.0 if bs_str.upper() == "FREE" else (_parse_price_float(bs_str) or 0.0)
            total = f"${bp + bs:.2f}"
        else:
            total = ""
        shipping_h = config.HEADER_MAP.get("buybox_shipping", "buybox_shipping")
        idx = headers.index(shipping_h) + 1 if shipping_h in headers else len(row)
        row.insert(idx, total)
    return row


@app.get("/api/export/fields")
async def api_export_fields():
    """返回可导出字段列表"""
    return {
        "fields": EXPORTABLE_FIELDS,
        "headers": {f: config.HEADER_MAP.get(f, f) for f in EXPORTABLE_FIELDS},
    }


@app.get("/api/export/all")
async def api_export_all(format: str = "xlsx", change_filter: str = "all", fields: str = None):
    selected = _parse_selected_fields(fields)
    name = f"all_{_cn_now().strftime('%Y%m%d_%H%M%S')}"
    if format == "csv":
        return await _export_csv_streaming(name, batch_id=None, change_filter=change_filter, selected_fields=selected)
    else:
        return await _export_xlsx_streaming(name, batch_id=None, change_filter=change_filter, selected_fields=selected)


@app.get("/api/export/{batch_name}")
async def api_export_batch(batch_name: str, format: str = "xlsx", change_filter: str = "all", fields: str = None):
    batch = await db.get_batch_by_name(batch_name)
    if not batch:
        raise HTTPException(404, f"批次不存在: {batch_name}")
    selected = _parse_selected_fields(fields)
    if format == "csv":
        return await _export_csv_streaming(batch_name, batch_id=batch["id"], change_filter=change_filter, selected_fields=selected)
    else:
        return await _export_xlsx_streaming(batch_name, batch_id=batch["id"], change_filter=change_filter, selected_fields=selected)


async def _export_xlsx_streaming(filename: str, batch_id: int = None,
                                  change_filter: str = "all", selected_fields=None):
    """write_only 模式 + 临时文件 + 流式响应（百万级不 OOM）"""
    import tempfile
    headers, field_keys, include_total = _get_export_headers(selected_fields)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="采集结果")
    ws.append(headers)

    count = 0
    async for item in db.iter_results(batch_id, change_filter=change_filter):
        ws.append(_prepare_row(item, field_keys, headers, include_total))
        count += 1

    if count == 0:
        wb.close()
        raise HTTPException(404, "无数据")

    os.makedirs(config.EXPORT_DIR, exist_ok=True)
    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix=".xlsx", prefix="export_",
        dir=config.EXPORT_DIR)
    tmp_path = tmp.name
    tmp.close()
    try:
        wb.save(tmp_path)
    except Exception:
        wb.close()
        os.unlink(tmp_path)
        raise
    wb.close()

    async def stream_and_cleanup():
        try:
            with open(tmp_path, "rb") as f:
                while True:
                    chunk = f.read(65536)
                    if not chunk:
                        break
                    yield chunk
        finally:
            os.unlink(tmp_path)

    safe = re.sub(r'[^a-zA-Z0-9_\-]', '_', filename)
    return StreamingResponse(
        stream_and_cleanup(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe}.xlsx"},
    )


async def _export_csv_streaming(filename: str, batch_id: int = None,
                                 change_filter: str = "all", selected_fields=None):
    """逐行 yield 的真流式 CSV（百万级不 OOM）"""
    headers, field_keys, include_total = _get_export_headers(selected_fields)

    async def generate():
        out = io.StringIO()
        csv.writer(out).writerow(headers)
        yield out.getvalue().encode("utf-8-sig")

        async for item in db.iter_results(batch_id, change_filter=change_filter):
            out = io.StringIO()
            csv.writer(out).writerow(_prepare_row(item, field_keys, headers, include_total))
            yield out.getvalue().encode("utf-8")

    safe = re.sub(r'[^a-zA-Z0-9_\-]', '_', filename)
    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={safe}.csv"},
    )


@app.get("/api/export/{batch_name}/screenshots")
async def api_export_screenshots(batch_name: str):
    ss_dir = os.path.join(config.SCREENSHOT_DIR, batch_name)
    if not os.path.isdir(ss_dir):
        raise HTTPException(404, "无截图文件")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in os.listdir(ss_dir):
            if fname.endswith(".png"):
                zf.write(os.path.join(ss_dir, fname), fname)

    output.seek(0)
    filename = f"screenshots_{batch_name}.zip"
    return StreamingResponse(
        output,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==================== API: 诊断 ====================

@app.get("/api/diagnostic")
async def api_diagnostic():
    total_asins = await db.get_total_asins()
    progress = await db.get_progress()
    return {
        "total_asins": total_asins,
        "task_progress": progress,
        "active_workers": len(_worker_registry),
        "settings_version": _settings_version,
    }


@app.post("/api/results/delete-by-file")
async def api_delete_by_file(file: UploadFile = File(...)):
    """上传文件识别 ASIN 后删除对应数据"""
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"文件过大：{len(content)//1024//1024}MB，上限 {MAX_UPLOAD_BYTES//1024//1024}MB")
    filename = file.filename or ""

    asins = []
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                for cell in row:
                    if cell:
                        val = str(cell).strip().upper()
                        if re.match(r'^B[0-9A-Z]{9}$', val):
                            asins.append(val)
        finally:
            wb.close()
    elif filename.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            for cell in row:
                val = cell.strip().upper()
                if re.match(r'^B[0-9A-Z]{9}$', val):
                    asins.append(val)
    else:
        text = content.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            val = line.strip().upper()
            if re.match(r'^B[0-9A-Z]{9}$', val):
                asins.append(val)

    # 去重
    asin_list = list(dict.fromkeys(asins))
    if not asin_list:
        raise HTTPException(400, "文件中未找到有效 ASIN")

    CHUNK = 500
    # 收集截图文件
    screenshot_files = []
    for i in range(0, len(asin_list), CHUNK):
        chunk = asin_list[i:i+CHUNK]
        placeholders = ",".join("?" * len(chunk))
        async with db._db.execute(
            f"SELECT file_path FROM screenshots WHERE asin IN ({placeholders}) AND file_path IS NOT NULL", chunk
        ) as c:
            screenshot_files.extend(row["file_path"] for row in await c.fetchall())

    # 分批删除
    async with db._write_lock:
        await db._db.execute("BEGIN")
        for i in range(0, len(asin_list), CHUNK):
            chunk = asin_list[i:i+CHUNK]
            placeholders = ",".join("?" * len(chunk))
            await db._db.execute(f"DELETE FROM asin_changes WHERE asin IN ({placeholders})", chunk)
            await db._db.execute(f"DELETE FROM screenshots WHERE asin IN ({placeholders})", chunk)
            await db._db.execute(f"DELETE FROM batch_asins WHERE asin IN ({placeholders})", chunk)
            await db._db.execute(f"DELETE FROM asin_data WHERE asin IN ({placeholders})", chunk)
        await db._db.execute("COMMIT")

    _remove_screenshot_files(screenshot_files)
    return {"ok": True, "deleted": len(asin_list), "asin_count": len(asin_list)}


@app.delete("/api/results")
async def api_delete_results(request: Request):
    """按条件删除采集结果"""
    body = await request.json()
    batch_id = body.get("batch_id")
    asins = body.get("asins")  # list of ASIN strings
    search = body.get("search")  # fuzzy search string

    # 构建 ASIN 列表
    target_asins = set()
    has_explicit_filter = bool(asins or search)

    if asins:
        # asins 列表也做上限保护，避免一次投递百万级 ASIN 触发巨量 SELECT
        if len(asins) > 100000:
            raise HTTPException(400, f"asins 列表过长（{len(asins)}），单次上限 100000")
        target_asins.update(asins)

    if search:
        # 限长防 DoS：500 字符 / 10 关键词 / 单词 100 字符
        search = str(search)[:500]
        terms = [t.strip()[:100] for t in search.split(",") if t.strip()][:10]
        if terms:
            or_clauses = []
            params = []
            for term in terms:
                or_clauses.append("(d.asin LIKE ? OR d.title LIKE ? OR d.brand LIKE ?)")
                params.extend([f"%{term}%", f"%{term}%", f"%{term}%"])
            where = " OR ".join(or_clauses)
            sql = f"SELECT d.asin FROM asin_data d WHERE {where}"
            async with db._db.execute(sql, params) as c:
                rows = await c.fetchall()
                for row in rows:
                    target_asins.add(row["asin"])

    # 纯 batch_id（无 asins/search）→ 删除该批次所有 ASIN
    if batch_id and not has_explicit_filter:
        sql = "SELECT asin FROM batch_asins WHERE batch_id = ?"
        async with db._db.execute(sql, (batch_id,)) as c:
            target_asins = {row["asin"] for row in await c.fetchall()}
    # batch_id + 其他条件 → 取交集
    elif batch_id and target_asins:
        sql = "SELECT asin FROM batch_asins WHERE batch_id = ?"
        async with db._db.execute(sql, (batch_id,)) as c:
            batch_asins = {row["asin"] for row in await c.fetchall()}
        target_asins &= batch_asins

    if not target_asins:
        return {"ok": True, "deleted": 0}

    asin_list = list(target_asins)
    CHUNK = 500  # SQLite 变量数上限安全值

    # 先收集截图物理文件路径
    screenshot_files = []
    for i in range(0, len(asin_list), CHUNK):
        chunk = asin_list[i:i+CHUNK]
        placeholders = ",".join("?" * len(chunk))
        async with db._db.execute(
            f"SELECT file_path FROM screenshots WHERE asin IN ({placeholders}) AND file_path IS NOT NULL", chunk
        ) as c:
            screenshot_files.extend(row["file_path"] for row in await c.fetchall())

    # 分批删除
    async with db._write_lock:
        await db._db.execute("BEGIN")
        for i in range(0, len(asin_list), CHUNK):
            chunk = asin_list[i:i+CHUNK]
            placeholders = ",".join("?" * len(chunk))
            await db._db.execute(f"DELETE FROM asin_changes WHERE asin IN ({placeholders})", chunk)
            await db._db.execute(f"DELETE FROM screenshots WHERE asin IN ({placeholders})", chunk)
            await db._db.execute(f"DELETE FROM batch_asins WHERE asin IN ({placeholders})", chunk)
            await db._db.execute(f"DELETE FROM asin_data WHERE asin IN ({placeholders})", chunk)
        await db._db.execute("COMMIT")

    # 删除物理截图文件
    _remove_screenshot_files(screenshot_files)
    return {"ok": True, "deleted": len(asin_list)}


# ==================== 定时采集管理 ====================

_SCHEDULES_DIR = os.path.join(config.PROJECT_DIR, "data", "schedules")


def _get_schedules() -> list:
    return _runtime_settings.get("auto_scrape_schedules", [])


def _save_schedules(schedules: list):
    _runtime_settings["auto_scrape_schedules"] = schedules
    _save_settings()


def _extract_asins_from_file(filepath: str) -> list:
    """从文件提取 ASIN 列表"""
    asins = []
    seen = set()
    if not os.path.isfile(filepath):
        return asins
    with open(filepath, "rb") as f:
        content = f.read()
    filename = filepath.lower()
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                for cell in row:
                    if cell:
                        val = str(cell).strip().upper()
                        if re.match(r'^B[0-9A-Z]{9}$', val) and val not in seen:
                            asins.append(val)
                            seen.add(val)
        finally:
            wb.close()
    elif filename.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            for cell in row:
                val = cell.strip().upper()
                if re.match(r'^B[0-9A-Z]{9}$', val) and val not in seen:
                    asins.append(val)
                    seen.add(val)
    else:
        text = content.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            val = line.strip().upper()
            if re.match(r'^B[0-9A-Z]{9}$', val) and val not in seen:
                asins.append(val)
                seen.add(val)
    return asins


@app.get("/api/schedules")
async def api_list_schedules():
    return {"schedules": _get_schedules()}


@app.post("/api/schedules")
async def api_create_schedule(request: Request,
                              file: UploadFile = File(...),
                              name: str = Form(""),
                              time_str: str = Form(..., alias="time"),
                              interval_days: int = Form(1),
                              needs_screenshot: bool = Form(False)):
    """创建定时采集任务"""
    # 验证时间格式
    try:
        h, m = map(int, time_str.split(":"))
        if not (0 <= h <= 23 and 0 <= m <= 59):
            raise ValueError
    except ValueError:
        raise HTTPException(400, "时间格式错误，应为 HH:MM")

    if interval_days < 1:
        raise HTTPException(400, "间隔天数至少为 1")

    # 保存 ASIN 文件
    os.makedirs(_SCHEDULES_DIR, exist_ok=True)
    import uuid
    sched_id = f"sched_{uuid.uuid4().hex[:8]}"
    ext = os.path.splitext(file.filename or "")[1] or ".txt"
    source_file = os.path.join(_SCHEDULES_DIR, f"{sched_id}{ext}")
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"文件过大：{len(content)//1024//1024}MB，上限 {MAX_UPLOAD_BYTES//1024//1024}MB")
    with open(source_file, "wb") as f:
        f.write(content)

    # 验证文件中有 ASIN
    asin_list = _extract_asins_from_file(source_file)
    if not asin_list:
        os.remove(source_file)
        raise HTTPException(400, "文件中未找到有效 ASIN")

    # 首次创建：last_run_date 设为昨天，确保首次检查时立即触发
    from datetime import timedelta
    yesterday = (_cn_now() - timedelta(days=interval_days)).strftime("%Y-%m-%d")

    sched = {
        "id": sched_id,
        "name": name or f"定时任务-{time_str}",
        "time": time_str,
        "interval_days": interval_days,
        "source_file": source_file,
        "asin_count": len(asin_list),
        "needs_screenshot": needs_screenshot,
        "enabled": True,
        "last_run_date": yesterday,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    schedules = _get_schedules()
    schedules.append(sched)
    _save_schedules(schedules)

    return {"ok": True, "schedule": sched, "schedules": schedules}


@app.put("/api/schedules/{sched_id}")
async def api_update_schedule(sched_id: str, request: Request):
    """修改定时任务（enabled/time/interval_days/name）"""
    body = await request.json()
    schedules = _get_schedules()
    target = None
    for s in schedules:
        if s.get("id") == sched_id:
            target = s
            break
    if target is None:
        raise HTTPException(404, "定时任务不存在")

    if "enabled" in body:
        target["enabled"] = bool(body["enabled"])
    if "name" in body:
        target["name"] = body["name"]
    if "time" in body:
        try:
            h, m = map(int, body["time"].split(":"))
            if 0 <= h <= 23 and 0 <= m <= 59:
                target["time"] = body["time"]
        except (ValueError, AttributeError):
            pass
    if "interval_days" in body:
        val = int(body["interval_days"])
        if val >= 1:
            target["interval_days"] = val

    _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


@app.delete("/api/schedules/{sched_id}")
async def api_delete_schedule(sched_id: str):
    """删除定时任务 + ASIN 文件"""
    schedules = _get_schedules()
    target = None
    new_schedules = []
    for s in schedules:
        if s.get("id") == sched_id:
            target = s
        else:
            new_schedules.append(s)
    if target is None:
        raise HTTPException(404, "定时任务不存在")

    # 删除 ASIN 文件
    source_file = target.get("source_file", "")
    if source_file and os.path.isfile(source_file):
        os.remove(source_file)

    _save_schedules(new_schedules)
    return {"ok": True, "schedules": new_schedules}


@app.post("/api/schedules/{sched_id}/run")
async def api_run_schedule_now(sched_id: str):
    """手动立即执行一次定时任务"""
    schedules = _get_schedules()
    target = None
    for s in schedules:
        if s.get("id") == sched_id:
            target = s
            break
    if target is None:
        raise HTTPException(404, "定时任务不存在")

    source_file = target.get("source_file", "")
    asin_list = _extract_asins_from_file(source_file)
    if not asin_list:
        raise HTTPException(400, "ASIN 文件为空或不存在")

    now = _cn_now()  # 批次名与 last_run 用中国时间
    batch_name = f"auto_{target.get('name', 'task')}_{now:%Y%m%d_%H%M}"
    zc = _runtime_settings.get("zip_code", config.DEFAULT_ZIP_CODE)
    ns = target.get("needs_screenshot", False)
    batch_id = await db.create_batch(batch_name, ns, is_auto=True)
    await db.create_tasks(batch_id, asin_list, zc, ns)

    target["last_run_date"] = now.strftime("%Y-%m-%d")
    _save_schedules(schedules)
    logger.info(f"手动执行定时任务: {batch_name}, {len(asin_list)} ASINs")

    return {"ok": True, "batch_id": batch_id, "batch_name": batch_name, "asin_count": len(asin_list)}


# ==================== 兼容旧 schedule API（settings.html 旧 UI 调用） ====================

@app.get("/api/auto-scrape/schedules")
async def api_legacy_list_schedules():
    return {"schedules": _get_schedules()}


@app.post("/api/auto-scrape/schedules")
async def api_legacy_add_schedule(request: Request):
    """旧式简单定时（无文件，使用全库 ASIN）"""
    body = await request.json()
    time_str = body.get("time", "")
    try:
        h, m = map(int, time_str.split(":"))
        if not (0 <= h <= 23 and 0 <= m <= 59):
            raise ValueError
    except ValueError:
        raise HTTPException(400, "时间格式错误")

    import uuid
    sched = {
        "id": f"sched_{uuid.uuid4().hex[:8]}",
        "name": f"全库采集-{time_str}",
        "time": time_str,
        "interval_days": 1,
        "source_file": "",  # 空=全库 ASIN
        "asin_count": 0,
        "needs_screenshot": False,
        "enabled": True,
        "last_run_date": "",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    schedules = _get_schedules()
    schedules.append(sched)
    _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


@app.put("/api/auto-scrape/schedules/{index}")
async def api_legacy_toggle_schedule(index: int, request: Request):
    body = await request.json()
    schedules = _get_schedules()
    if 0 <= index < len(schedules):
        if "enabled" in body:
            schedules[index]["enabled"] = bool(body["enabled"])
        _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


@app.delete("/api/auto-scrape/schedules/{index}")
async def api_legacy_delete_schedule(index: int):
    schedules = _get_schedules()
    if 0 <= index < len(schedules):
        removed = schedules.pop(index)
        sf = removed.get("source_file", "")
        if sf and os.path.isfile(sf):
            os.remove(sf)
        _save_schedules(schedules)
    return {"ok": True, "schedules": schedules}


# ==================== Recon 侦查端点（锁竞争 / 阶段耗时）====================
# 临时仪表，用于诊断 _write_lock 竞争和 accept_results_batch 内部瓶颈
# 完成诊断后可删除（含 common/database.py 中的 TimedLock / LOCK_STATS）

def _pct(samples: list, p: float) -> float:
    if not samples:
        return 0.0
    s = sorted(samples)
    i = max(0, min(len(s) - 1, int(len(s) * p)))
    return round(s[i], 2)


def _summary(samples: list) -> dict:
    if not samples:
        return {"count": 0}
    return {
        "count": len(samples),
        "p50": _pct(samples, 0.50),
        "p95": _pct(samples, 0.95),
        "p99": _pct(samples, 0.99),
        "max": round(max(samples), 2),
        "mean": round(sum(samples) / len(samples), 2),
    }


@app.get("/api/_debug/lock-stats")
async def api_debug_lock_stats():
    """返回锁等待 / 持锁 / 内部分阶段耗时分布。单位：毫秒。"""
    from common.database import LOCK_STATS
    return {
        "waits": {k: _summary(list(v)) for k, v in LOCK_STATS["waits"].items()},
        "holds": {k: _summary(list(v)) for k, v in LOCK_STATS["holds"].items()},
        "stage_timings": {k: _summary(list(v)) for k, v in LOCK_STATS["stage_timings"].items()},
        "slow_holds_recent": LOCK_STATS["slow_holds"][-50:],
        "slow_holds_count": len(LOCK_STATS["slow_holds"]),
    }


@app.post("/api/_debug/lock-stats/reset")
async def api_debug_lock_stats_reset():
    """清空所有计时统计（开始新一轮观察前调用）。"""
    from common.database import LOCK_STATS
    LOCK_STATS["waits"].clear()
    LOCK_STATS["holds"].clear()
    LOCK_STATS["stage_timings"].clear()
    LOCK_STATS["slow_holds"].clear()
    return {"ok": True}


@app.delete("/api/database")
async def api_clear_database():
    """清空所有数据 + 截图文件"""
    import shutil
    async with db._write_lock:
        await db._db.execute("BEGIN")
        for table in ["asin_changes", "asin_data", "batch_asins", "tasks", "screenshots", "batches"]:
            await db._db.execute(f"DELETE FROM {table}")
        await db._db.execute("DELETE FROM sqlite_sequence")
        await db._db.execute("COMMIT")

    # 清理截图文件
    ss_dir = config.SCREENSHOT_DIR
    if os.path.isdir(ss_dir):
        shutil.rmtree(ss_dir)
        os.makedirs(ss_dir, exist_ok=True)

    return {"ok": True}


# ==================== 入口 ====================

def main():
    import uvicorn
    uvicorn.run(
        "server.app:app",
        host=config.SERVER_HOST,
        port=config.SERVER_PORT,
        log_level="info",
    )


if __name__ == "__main__":
    main()
